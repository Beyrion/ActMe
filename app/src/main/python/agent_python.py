import builtins
import contextlib
import io
import json
import os
import py_compile
import sys
import time
import traceback
import ctypes
import multiprocessing
import subprocess
import pathlib
from os.path import abspath, commonpath, dirname, isabs, join, realpath


DENIED_IMPORT_ROOTS = {
    "ensurepip",
    "pip",
    "venv",
    "webbrowser",
}


def run_code(code, input_text="", timeout_ms=3000, workspace_dir="", report_font_path=""):
    start = time.monotonic()
    deadline = start + max(1, min(int(timeout_ms or 3000), 30000)) / 1000.0
    stdout = io.StringIO()
    stderr = io.StringIO()
    state = {"result": None, "has_result": False}
    original_cwd = None

    def check_deadline():
        if time.monotonic() > deadline:
            raise TimeoutError("python_exec exceeded timeout")

    def trace_func(frame, event, arg):
        check_deadline()
        return trace_func

    original_import = builtins.__import__
    import_state = {"depth": 0}

    def safe_import(name, globals=None, locals=None, fromlist=(), level=0):
        root = name.split(".", 1)[0]
        if root in DENIED_IMPORT_ROOTS:
            raise ImportError("module is not allowed: " + name)
        import_state["depth"] += 1
        try:
            return original_import(name, globals, locals, fromlist, level)
        finally:
            import_state["depth"] -= 1

    def set_result(value):
        state["result"] = value
        state["has_result"] = True

    workspace_root = realpath(workspace_dir) if workspace_dir else ""
    report_font_path = realpath(report_font_path) if report_font_path else ""
    runtime_roots = _runtime_roots()
    if workspace_root:
        os.environ.setdefault("HOME", workspace_root)
        os.environ.setdefault("MPLCONFIGDIR", join(workspace_root, ".matplotlib"))
        os.environ.setdefault("XDG_CACHE_HOME", join(workspace_root, ".cache"))
        if report_font_path:
            os.environ["ACTME_REPORT_FONT_PATH"] = report_font_path
        os.makedirs(os.environ["MPLCONFIGDIR"], exist_ok=True)
        os.makedirs(os.environ["XDG_CACHE_HOME"], exist_ok=True)
        original_cwd = os.getcwd()
        os.chdir(workspace_root)
    before_files = _workspace_snapshot(workspace_root)

    def validate_workspace_path(path):
        if not path:
            raise PermissionError("empty path")
        raw_path = str(path)
        if workspace_root and not isabs(raw_path):
            raw_path = join(workspace_root, raw_path)
        return realpath(abspath(raw_path))

    def resolve_read_path(path):
        if not path:
            raise PermissionError("empty path")
        raw_path = str(path)
        if workspace_root and not isabs(raw_path):
            raw_path = join(workspace_root, raw_path)
        return realpath(abspath(raw_path))

    def is_write_mode(mode):
        text = str(mode or "r")
        return any(flag in text for flag in ("w", "a", "x", "+"))

    def is_runtime_path(path):
        return import_state["depth"] > 0 and _is_under_any(path, runtime_roots)

    def safe_open(file, mode="r", *args, **kwargs):
        if is_write_mode(mode):
            if is_runtime_path(file):
                return open(file, mode, *args, **kwargs)
            return open(validate_workspace_path(file), mode, *args, **kwargs)
        return open(resolve_read_path(file), mode, *args, **kwargs)

    def read_excel(path, max_rows=200, max_sheets=10, values_only=True):
        full_path = resolve_read_path(path)
        return _read_excel(full_path, max_rows=max_rows, max_sheets=max_sheets, values_only=values_only)

    def write_excel(filename, sheets):
        full_path = validate_workspace_path(filename)
        return _write_excel(full_path, sheets)

    def write_report(markdown_text, base_name="report", title=None, make_pdf=True):
        return _write_report_bundle(
            markdown_text,
            base_name=base_name,
            title=title,
            make_pdf=make_pdf,
            validate_write_path=validate_workspace_path,
            workspace_root=workspace_root,
            report_font_path=report_font_path,
        )

    def script_path(name):
        safe_name = _safe_script_name(name)
        script_dir = validate_workspace_path("python")
        os.makedirs(script_dir, exist_ok=True)
        return validate_workspace_path(join("python", safe_name))

    def save_script(name, source):
        full_path = script_path(name)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        with open(full_path, "w", encoding="utf-8") as f:
            f.write(str(source or ""))
        return {"path": full_path, "bytes": len(str(source or "").encode("utf-8"))}

    def load_script(name):
        full_path = script_path(name)
        with open(full_path, "r", encoding="utf-8") as f:
            return {"path": full_path, "source": f.read()}

    def list_scripts():
        script_dir = validate_workspace_path("python")
        os.makedirs(script_dir, exist_ok=True)
        scripts = []
        for root, _, files in os.walk(script_dir):
            for filename in files:
                if not filename.endswith(".py"):
                    continue
                full_path = join(root, filename)
                rel = os.path.relpath(full_path, script_dir).replace("\\", "/")
                scripts.append({"name": rel, "path": full_path, "bytes": os.path.getsize(full_path)})
        return sorted(scripts, key=lambda item: item["name"])

    def run_script(name):
        full_path = script_path(name)
        with open(full_path, "r", encoding="utf-8") as f:
            source = f.read()
        compiled_script = compile(source, full_path, "exec")
        exec(compiled_script, globals_dict, globals_dict)
        return state["result"] if state["has_result"] else globals_dict.get("result", None)

    def compile_script(name):
        full_path = script_path(name)
        try:
            py_compile.compile(full_path, doraise=True)
            return {"ok": True, "path": full_path, "error": ""}
        except py_compile.PyCompileError as exc:
            return {"ok": False, "path": full_path, "error": str(exc)}

    safe_builtins = {
        "abs": abs,
        "all": all,
        "any": any,
        "bool": bool,
        "bytes": bytes,
        "chr": chr,
        "dict": dict,
        "divmod": divmod,
        "enumerate": enumerate,
        "filter": filter,
        "float": float,
        "format": format,
        "hash": hash,
        "hex": hex,
        "int": int,
        "isinstance": isinstance,
        "issubclass": issubclass,
        "len": len,
        "list": list,
        "map": map,
        "max": max,
        "min": min,
        "next": next,
        "ord": ord,
        "pow": pow,
        "print": print,
        "range": range,
        "repr": repr,
        "reversed": reversed,
        "round": round,
        "set": set,
        "slice": slice,
        "sorted": sorted,
        "str": str,
        "sum": sum,
        "tuple": tuple,
        "zip": zip,
        "Exception": Exception,
        "ValueError": ValueError,
        "TypeError": TypeError,
        "RuntimeError": RuntimeError,
        "TimeoutError": TimeoutError,
        "OSError": OSError,
        "PermissionError": PermissionError,
        "FileNotFoundError": FileNotFoundError,
        "ImportError": ImportError,
        "KeyError": KeyError,
        "IndexError": IndexError,
        "AttributeError": AttributeError,
        "ZeroDivisionError": ZeroDivisionError,
        "__build_class__": builtins.__build_class__,
        "__import__": safe_import,
        "open": safe_open,
    }

    globals_dict = {
        "__builtins__": safe_builtins,
        "__name__": "__agent_python__",
        "input_text": input_text or "",
        "workspace_dir": workspace_root,
        "report_font_dir": report_font_path,
        "read_excel": read_excel,
        "write_excel": write_excel,
        "write_report": write_report,
        "save_script": save_script,
        "load_script": load_script,
        "list_scripts": list_scripts,
        "run_script": run_script,
        "compile_script": compile_script,
        "emit": set_result,
        "set_result": set_result,
    }

    try:
        globals_dict["input_json"] = json.loads(input_text) if input_text else None
    except Exception:
        globals_dict["input_json"] = None

    ok = True
    error = ""
    restore_os = _patch_os_for_sandbox(workspace_root, validate_workspace_path, is_runtime_path)
    restore_io = _patch_io_for_sandbox(validate_workspace_path, resolve_read_path, is_write_mode, is_runtime_path)
    restore_pathlib = _patch_pathlib_for_sandbox(validate_workspace_path, resolve_read_path, is_write_mode, is_runtime_path)
    restore_system_modules = _patch_system_modules_for_sandbox()
    sys.settrace(trace_func)
    try:
        compiled = compile(code or "", "<agent_python>", "exec")
        with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
            exec(compiled, globals_dict, globals_dict)
    except Exception:
        ok = False
        error = traceback.format_exc()
    finally:
        sys.settrace(None)
        restore_system_modules()
        restore_pathlib()
        restore_io()
        restore_os()
        if original_cwd:
            os.chdir(original_cwd)

    result = state["result"] if state["has_result"] else globals_dict.get("result", None)
    output_files = _workspace_changed_files(workspace_root, before_files)
    return json.dumps(
        {
            "ok": ok,
            "stdout": stdout.getvalue(),
            "stderr": stderr.getvalue(),
            "error": error,
            "result": _json_safe(result),
            "elapsed_ms": int((time.monotonic() - start) * 1000),
            "output_files": output_files,
        },
        ensure_ascii=False,
    )


def _runtime_roots():
    roots = set()
    for path in list(sys.path) + [__file__]:
        if not path:
            continue
        full = realpath(abspath(path))
        parts = full.split(os.sep)
        if "AssetFinder" in parts:
            idx = parts.index("AssetFinder")
            roots.add(os.sep.join(parts[: idx + 1]))
        elif "chaquopy" in parts:
            idx = parts.index("chaquopy")
            roots.add(os.sep.join(parts[: idx + 1]))
        else:
            roots.add(dirname(full) if os.path.isfile(full) else full)
    return sorted(roots, key=len, reverse=True)


def _is_under_any(path, roots):
    if not roots:
        return False
    try:
        full = realpath(abspath(str(path)))
    except Exception:
        return False
    for root in roots:
        try:
            if commonpath([root, full]) == root:
                return True
        except Exception:
            continue
    return False


def _patch_os_for_sandbox(workspace_root, validate_workspace_path, is_runtime_path):
    patched = {}

    def remember(name):
        if hasattr(os, name):
            patched[name] = getattr(os, name)

    def deny(name):
        def denied(*args, **kwargs):
            raise PermissionError("os." + name + " is not allowed in python sandbox")
        return denied

    def workspace_path_call(name):
        original = getattr(os, name)

        def wrapped(path, *args, **kwargs):
            full = realpath(abspath(str(path)))
            if is_runtime_path(full):
                return original(path, *args, **kwargs)
            if name in ("makedirs", "mkdir"):
                if workspace_root and commonpath([workspace_root, full]) != workspace_root and os.path.isdir(full):
                    return None
            return original(validate_workspace_path(path), *args, **kwargs)

        return wrapped

    def workspace_two_path_call(name):
        original = getattr(os, name)

        def wrapped(src, dst, *args, **kwargs):
            if is_runtime_path(src) and is_runtime_path(dst):
                return original(src, dst, *args, **kwargs)
            return original(validate_workspace_path(src), validate_workspace_path(dst), *args, **kwargs)

        return wrapped

    denied_names = [
        "chdir",
        "chmod",
        "chown",
        "execv",
        "execve",
        "fork",
        "kill",
        "killpg",
        "popen",
        "spawnl",
        "spawnle",
        "spawnlp",
        "spawnlpe",
        "spawnv",
        "spawnve",
        "spawnvp",
        "spawnvpe",
        "startfile",
        "system",
    ]
    one_path_names = ["makedirs", "mkdir", "remove", "rmdir", "unlink"]
    two_path_names = ["replace", "rename"]

    for name in denied_names + one_path_names + two_path_names:
        remember(name)
    for name in denied_names:
        if hasattr(os, name):
            setattr(os, name, deny(name))
    for name in one_path_names:
        if hasattr(os, name):
            setattr(os, name, workspace_path_call(name))
    for name in two_path_names:
        if hasattr(os, name):
            setattr(os, name, workspace_two_path_call(name))

    def restore():
        for name, original in patched.items():
            setattr(os, name, original)

    return restore


def _patch_io_for_sandbox(validate_workspace_path, resolve_read_path, is_write_mode, is_runtime_path):
    original_open = io.open

    def safe_io_open(file, mode="r", *args, **kwargs):
        if isinstance(file, int):
            return original_open(file, mode, *args, **kwargs)
        if is_write_mode(mode):
            if is_runtime_path(file):
                return original_open(file, mode, *args, **kwargs)
            return original_open(validate_workspace_path(file), mode, *args, **kwargs)
        return original_open(resolve_read_path(file), mode, *args, **kwargs)

    io.open = safe_io_open

    def restore():
        io.open = original_open

    return restore


def _patch_pathlib_for_sandbox(validate_workspace_path, resolve_read_path, is_write_mode, is_runtime_path):
    original_open = pathlib.Path.open
    original_read_text = pathlib.Path.read_text
    original_read_bytes = pathlib.Path.read_bytes
    original_write_text = pathlib.Path.write_text
    original_write_bytes = pathlib.Path.write_bytes

    def target_for(path, mode):
        raw = str(path)
        if is_write_mode(mode):
            if is_runtime_path(raw):
                return pathlib.Path(raw)
            return pathlib.Path(validate_workspace_path(raw))
        return pathlib.Path(resolve_read_path(raw))

    def safe_path_open(self, mode="r", buffering=-1, encoding=None, errors=None, newline=None):
        return original_open(target_for(self, mode), mode, buffering, encoding, errors, newline)

    def safe_read_text(self, encoding=None, errors=None):
        return original_read_text(target_for(self, "r"), encoding=encoding, errors=errors)

    def safe_read_bytes(self):
        return original_read_bytes(target_for(self, "rb"))

    def safe_write_text(self, data, encoding=None, errors=None, newline=None):
        return original_write_text(target_for(self, "w"), data, encoding=encoding, errors=errors, newline=newline)

    def safe_write_bytes(self, data):
        return original_write_bytes(target_for(self, "wb"), data)

    pathlib.Path.open = safe_path_open
    pathlib.Path.read_text = safe_read_text
    pathlib.Path.read_bytes = safe_read_bytes
    pathlib.Path.write_text = safe_write_text
    pathlib.Path.write_bytes = safe_write_bytes

    def restore():
        pathlib.Path.open = original_open
        pathlib.Path.read_text = original_read_text
        pathlib.Path.read_bytes = original_read_bytes
        pathlib.Path.write_text = original_write_text
        pathlib.Path.write_bytes = original_write_bytes

    return restore


def _patch_system_modules_for_sandbox():
    patched = []

    def remember(module, name):
        if hasattr(module, name):
            patched.append((module, name, getattr(module, name)))

    def deny(label):
        def denied(*args, **kwargs):
            raise PermissionError(label + " is not allowed in python sandbox")
        return denied

    for module, names in [
        (subprocess, ["Popen", "call", "check_call", "check_output", "run"]),
        (ctypes, ["CDLL", "PyDLL", "WinDLL", "OleDLL", "LibraryLoader"]),
        (multiprocessing, ["Process", "Pool", "Manager"]),
    ]:
        for name in names:
            remember(module, name)
            if hasattr(module, name):
                setattr(module, name, deny(module.__name__ + "." + name))

    def restore():
        for module, name, original in patched:
            setattr(module, name, original)

    return restore


def _workspace_snapshot(workspace_root):
    if not workspace_root or not os.path.isdir(workspace_root):
        return {}
    snapshot = {}
    for root, _, files in os.walk(workspace_root):
        for filename in files:
            full = join(root, filename)
            try:
                rel = os.path.relpath(full, workspace_root).replace("\\", "/")
                snapshot[rel] = (os.path.getmtime(full), os.path.getsize(full))
            except Exception:
                pass
    return snapshot


def _workspace_changed_files(workspace_root, before):
    if not workspace_root or not os.path.isdir(workspace_root):
        return []
    after = _workspace_snapshot(workspace_root)
    changed = []
    for rel, stat in after.items():
        if rel.startswith("python/"):
            continue
        if before.get(rel) != stat:
            changed.append(rel)
    return sorted(changed)


def _json_safe(value):
    try:
        json.dumps(value)
        return value
    except Exception:
        return repr(value)


def _read_excel(path, max_rows=200, max_sheets=10, values_only=True):
    from datetime import date, datetime, time as dt_time
    from decimal import Decimal
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for sheet_name in workbook.sheetnames[: max(1, int(max_sheets or 10))]:
            ws = workbook[sheet_name]
            rows = []
            for row_index, row in enumerate(ws.iter_rows(values_only=values_only), start=1):
                if row_index > max(1, int(max_rows or 200)):
                    break
                rows.append([_cell_value(v) for v in row])
            sheets.append(
                {
                    "name": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "rows_returned": len(rows),
                    "rows": rows,
                }
            )
        return {"path": path, "sheet_count": len(workbook.sheetnames), "sheets": sheets}
    finally:
        workbook.close()


def _cell_value(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _safe_script_name(name):
    raw = str(name or "").strip().replace("\\", "/")
    if not raw:
        raise ValueError("script name is empty")
    parts = [part for part in raw.split("/") if part not in ("", ".", "..")]
    safe_parts = []
    for part in parts:
        cleaned = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in part)
        if cleaned:
            safe_parts.append(cleaned)
    if not safe_parts:
        raise ValueError("invalid script name")
    filename = "/".join(safe_parts)
    if not filename.endswith(".py"):
        filename += ".py"
    return filename


def _safe_output_path(name, default_name="report", default_ext=".md"):
    raw = str(name or default_name).strip().replace("\\", "/")
    if not raw:
        raw = default_name
    raw = raw.split("?", 1)[0].split("#", 1)[0]
    raw = raw.lstrip("/")
    if ":" in raw.split("/", 1)[0]:
        raw = raw.split(":", 1)[-1].lstrip("/")
    parts = []
    for part in raw.split("/"):
        if part in ("", ".", ".."):
            continue
        cleaned = "".join(ch if ch.isalnum() or ch in " ._()-[]" else "_" for ch in part).strip()
        if cleaned:
            parts.append(cleaned)
    if not parts:
        parts = [default_name]
    rel = "/".join(parts)
    if default_ext and not rel.lower().endswith(default_ext.lower()):
        rel += default_ext
    return rel


def _replace_ext(path, ext):
    root = path.rsplit(".", 1)[0] if "." in path.rsplit("/", 1)[-1] else path
    return root + ext


def _write_text_file(path, text):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    # Use BOM for report text artifacts so Windows editors and Markdown viewers
    # reliably detect UTF-8 Chinese content.
    with open(path, "w", encoding="utf-8-sig") as f:
        f.write(text)


def _write_report_bundle(markdown_text, base_name="report", title=None, make_pdf=True, validate_write_path=None, workspace_root="", report_font_path=""):
    markdown = str(markdown_text or "").strip()
    if not markdown:
        raise ValueError("markdown_text is empty")
    md_rel = _safe_output_path(base_name, default_name="report", default_ext=".md")
    html_rel = _replace_ext(md_rel, ".html")
    pdf_rel = _replace_ext(md_rel, ".pdf")

    md_path = validate_write_path(md_rel) if validate_write_path else realpath(abspath(md_rel))
    html_path = validate_write_path(html_rel) if validate_write_path else realpath(abspath(html_rel))
    pdf_path = validate_write_path(pdf_rel) if validate_write_path else realpath(abspath(pdf_rel))

    _write_text_file(md_path, markdown)
    html = _markdown_to_html_document(markdown, title=title)
    _write_text_file(html_path, html)
    print("[write_report] markdown_written path={} exists={} bytes={}".format(md_path, os.path.exists(md_path), os.path.getsize(md_path) if os.path.exists(md_path) else -1))
    print("[write_report] html_written path={} exists={} bytes={}".format(html_path, os.path.exists(html_path), os.path.getsize(html_path) if os.path.exists(html_path) else -1))

    output_files = [md_rel, html_rel]
    result = {
        "markdown": md_rel,
        "html": html_rel,
        "pdf": None,
        "output_files": list(output_files),
    }
    if make_pdf:
        print("[write_report] pdf_start path={} title={}".format(pdf_path, title or ""))
        try:
            print("[write_report] pdf_html_start path={} font={}".format(pdf_path, report_font_path or ""))
            _html_to_pdf_fpdf2(html, pdf_path, font_path=report_font_path)
            print("[write_report] pdf_html_success path={} exists={} bytes={}".format(pdf_path, os.path.exists(pdf_path), os.path.getsize(pdf_path) if os.path.exists(pdf_path) else -1))
        except Exception as html_exc:
            result["pdf_html_error"] = str(html_exc)
            result["pdf_html_error_trace"] = traceback.format_exc()
            print("[write_report] pdf_html_error path={} exists={} bytes={}".format(pdf_path, os.path.exists(pdf_path), os.path.getsize(pdf_path) if os.path.exists(pdf_path) else -1), file=sys.stderr)
            print(result["pdf_html_error_trace"], file=sys.stderr)
            try:
                print("[write_report] pdf_reportlab_start path={} font={}".format(pdf_path, report_font_path or ""))
                _markdown_to_pdf_reportlab(markdown, pdf_path, title=title, font_path=report_font_path)
                print("[write_report] pdf_reportlab_success path={} exists={} bytes={}".format(pdf_path, os.path.exists(pdf_path), os.path.getsize(pdf_path) if os.path.exists(pdf_path) else -1))
            except Exception:
                raise
        try:
            print("[write_report] pdf_success path={} exists={} bytes={}".format(pdf_path, os.path.exists(pdf_path), os.path.getsize(pdf_path) if os.path.exists(pdf_path) else -1))
            output_files.append(pdf_rel)
            result["pdf"] = pdf_rel
            result["output_files"] = list(output_files)
        except Exception as exc:
            result["pdf_error"] = str(exc)
            result["pdf_error_trace"] = traceback.format_exc()
            print("[write_report] pdf_error path={} exists={} bytes={}".format(pdf_path, os.path.exists(pdf_path), os.path.getsize(pdf_path) if os.path.exists(pdf_path) else -1), file=sys.stderr)
            print(result["pdf_error_trace"], file=sys.stderr)
    else:
        print("[write_report] pdf_skipped make_pdf=false")
    return result


def _html_to_pdf_fpdf2(html_text, pdf_path, font_path=""):
    from fpdf import FPDF

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    font_family = "Helvetica"
    font_errors = []
    for candidate in _report_font_candidates(font_path):
        try:
            font_family = "ActMeReportFont"
            for style in ("", "B", "I", "BI"):
                pdf.add_font(font_family, style=style, fname=candidate)
            print("[write_report] fpdf_font_ok path={}".format(candidate))
            break
        except Exception as exc:
            font_errors.append("{}: {}".format(candidate, exc))
            print("[write_report] fpdf_font_error path={} error={}".format(candidate, exc), file=sys.stderr)
            font_family = "Helvetica"
    if font_family == "Helvetica" and font_errors:
        print("[write_report] fpdf_font_fallback errors={}".format(" | ".join(font_errors)), file=sys.stderr)
    pdf.set_font(font_family, size=11)
    body_html = _html_body_for_fpdf2(html_text)
    pdf.write_html(body_html)
    pdf.output(pdf_path)


def _html_body_for_fpdf2(html_text):
    import re
    text = re.sub(r"(?is)<style[^>]*>.*?</style>", "", html_text)
    text = re.sub(r"(?is)<script[^>]*>.*?</script>", "", text)
    match = re.search(r"(?is)<body[^>]*>(.*?)</body>", text)
    if match:
        text = match.group(1)
    text = re.sub(r"(?is)<pre[^>]*><code[^>]*>(.*?)</code></pre>", r"<pre>\1</pre>", text)
    text = re.sub(r"(?is)<code[^>]*>(.*?)</code>", r"<font face=\"Courier\">\1</font>", text)
    text = re.sub(r"(?is)<hr\s*/?>", "<br>", text)
    return text


def _report_font_candidates(font_path=""):
    candidates = []
    if font_path:
        if os.path.isdir(font_path):
            for name in (
                "ActMeReportSans.ttf",
                "ActMeReportSerif.ttf",
            ):
                candidates.append(os.path.join(font_path, name))
        else:
            candidates.append(font_path)
    env_path = os.environ.get("ACTME_REPORT_FONT_PATH", "")
    if env_path and env_path != font_path:
        if os.path.isdir(env_path):
            for name in (
                "ActMeReportSans.ttf",
                "ActMeReportSerif.ttf",
            ):
                candidates.append(os.path.join(env_path, name))
        else:
            candidates.append(env_path)
    return [path for path in dict.fromkeys(candidates) if path and os.path.exists(path)]


def _markdown_to_html_document(markdown_text, title=None):
    body = None
    try:
        import markdown as markdown_pkg
        body = markdown_pkg.markdown(
            markdown_text,
            extensions=["extra", "tables", "fenced_code", "toc"],
            output_format="html5",
        )
    except Exception:
        body = _markdown_to_html_fallback(markdown_text)
    page_title = _html_escape(title or _first_markdown_heading(markdown_text) or "Report")
    return """<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{title}</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Noto Sans CJK SC", "Microsoft YaHei", sans-serif; line-height: 1.65; max-width: 920px; margin: 32px auto; padding: 0 20px; color: #1f2933; }}
    h1, h2, h3 {{ line-height: 1.25; color: #111827; }}
    h1 {{ border-bottom: 2px solid #e5e7eb; padding-bottom: 12px; }}
    table {{ border-collapse: collapse; width: 100%; margin: 16px 0; }}
    th, td {{ border: 1px solid #d1d5db; padding: 8px 10px; vertical-align: top; }}
    th {{ background: #f3f4f6; }}
    code, pre {{ background: #f3f4f6; border-radius: 4px; }}
    pre {{ padding: 12px; overflow-x: auto; }}
    blockquote {{ border-left: 4px solid #d1d5db; margin-left: 0; padding-left: 14px; color: #4b5563; }}
  </style>
</head>
<body>
{body}
</body>
</html>
""".format(title=page_title, body=body)


def _markdown_to_html_fallback(markdown_text):
    lines = markdown_text.splitlines()
    html_parts = []
    paragraph = []
    in_code = False
    code_lines = []
    i = 0

    def flush_paragraph():
        if paragraph:
            html_parts.append("<p>" + "<br>\n".join(_inline_markdown_to_html(x) for x in paragraph) + "</p>")
            paragraph.clear()

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if stripped.startswith("```"):
            if in_code:
                html_parts.append("<pre><code>" + _html_escape("\n".join(code_lines)) + "</code></pre>")
                code_lines = []
                in_code = False
            else:
                flush_paragraph()
                in_code = True
            i += 1
            continue
        if in_code:
            code_lines.append(line)
            i += 1
            continue
        if not stripped:
            flush_paragraph()
            i += 1
            continue
        if _is_markdown_table(lines, i):
            flush_paragraph()
            table_len = _markdown_table_block_length(lines, i)
            html_parts.append(_markdown_table_to_html(lines[i : i + table_len]))
            i += table_len
            continue
        heading = _heading_level_and_text(stripped)
        if heading:
            flush_paragraph()
            level, text = heading
            html_parts.append("<h{0}>{1}</h{0}>".format(level, _inline_markdown_to_html(text)))
        elif stripped.startswith(("- ", "* ")):
            flush_paragraph()
            items = []
            while i < len(lines) and lines[i].strip().startswith(("- ", "* ")):
                items.append("<li>" + _inline_markdown_to_html(lines[i].strip()[2:].strip()) + "</li>")
                i += 1
            html_parts.append("<ul>" + "\n".join(items) + "</ul>")
            continue
        else:
            paragraph.append(line)
        i += 1
    flush_paragraph()
    if in_code:
        html_parts.append("<pre><code>" + _html_escape("\n".join(code_lines)) + "</code></pre>")
    return "\n".join(html_parts)


def _markdown_to_pdf_reportlab(markdown_text, pdf_path, title=None, font_path=""):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "ActMeReportFont"
    font_errors = []
    for candidate in _report_font_candidates(font_path):
        try:
            pdfmetrics.registerFont(TTFont(font_name, candidate))
            print("[write_report] reportlab_font_ok path={}".format(candidate))
            break
        except Exception as exc:
            font_errors.append("{}: {}".format(candidate, exc))
            print("[write_report] reportlab_font_error path={} error={}".format(candidate, exc), file=sys.stderr)
    else:
        font_name = "STSong-Light"
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(font_name))
        except Exception:
            if font_errors:
                print("[write_report] reportlab_font_fallback errors={}".format(" | ".join(font_errors)), file=sys.stderr)
            font_name = "Helvetica"

    styles = getSampleStyleSheet()
    base = ParagraphStyle(
        "ActMeBase",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=10.5,
        leading=16,
        spaceAfter=7,
    )
    h1 = ParagraphStyle("ActMeH1", parent=base, fontSize=20, leading=26, spaceBefore=8, spaceAfter=14)
    h2 = ParagraphStyle("ActMeH2", parent=base, fontSize=15, leading=21, spaceBefore=10, spaceAfter=8)
    h3 = ParagraphStyle("ActMeH3", parent=base, fontSize=12.5, leading=18, spaceBefore=8, spaceAfter=6)
    bullet = ParagraphStyle("ActMeBullet", parent=base, leftIndent=14, firstLineIndent=-8)

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=title or _first_markdown_heading(markdown_text) or "Report",
    )
    story = []
    paragraph = []
    lines = markdown_text.splitlines()
    i = 0

    def flush_paragraph():
        if paragraph:
            story.append(Paragraph(_inline_markdown_to_reportlab(" ".join(x.strip() for x in paragraph)), base))
            paragraph.clear()

    while i < len(lines):
        stripped = lines[i].strip()
        if not stripped:
            flush_paragraph()
            i += 1
            continue
        if stripped.startswith("```"):
            flush_paragraph()
            code = []
            i += 1
            while i < len(lines) and not lines[i].strip().startswith("```"):
                code.append(lines[i])
                i += 1
            story.append(Paragraph(_html_escape("\n".join(code)).replace("\n", "<br/>"), base))
            story.append(Spacer(1, 5))
            i += 1
            continue
        if _is_markdown_table(lines, i):
            flush_paragraph()
            table_len = _markdown_table_block_length(lines, i)
            table_rows = _parse_markdown_table(lines[i : i + table_len])
            data = [[Paragraph(_inline_markdown_to_reportlab(cell), base) for cell in row] for row in table_rows]
            table = Table(data, hAlign="LEFT", repeatRows=1)
            table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(table)
            story.append(Spacer(1, 8))
            i += table_len
            continue
        heading = _heading_level_and_text(stripped)
        if heading:
            flush_paragraph()
            level, text = heading
            story.append(Paragraph(_inline_markdown_to_reportlab(text), h1 if level == 1 else h2 if level == 2 else h3))
        elif stripped.startswith(("- ", "* ")):
            flush_paragraph()
            story.append(Paragraph("• " + _inline_markdown_to_reportlab(stripped[2:].strip()), bullet))
        else:
            paragraph.append(stripped)
        i += 1
    flush_paragraph()
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    doc.build(story)


def _html_escape(text):
    import html
    return html.escape(str(text or ""), quote=True)


def _inline_markdown_to_html(text):
    import re
    escaped = _html_escape(text)
    escaped = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", escaped)
    escaped = re.sub(r"`(.+?)`", r"<code>\1</code>", escaped)
    return escaped


def _inline_markdown_to_reportlab(text):
    import re
    escaped = _html_escape(text)
    escaped = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", escaped)
    escaped = re.sub(r"`(.+?)`", r"<font name='Courier'>\1</font>", escaped)
    return escaped


def _heading_level_and_text(stripped):
    if not stripped.startswith("#"):
        return None
    hashes = len(stripped) - len(stripped.lstrip("#"))
    if hashes < 1 or hashes > 6 or len(stripped) <= hashes or stripped[hashes] != " ":
        return None
    return min(hashes, 3), stripped[hashes:].strip()


def _first_markdown_heading(markdown_text):
    for line in str(markdown_text or "").splitlines():
        heading = _heading_level_and_text(line.strip())
        if heading:
            return heading[1]
    return None


def _is_markdown_table(lines, index):
    if index + 1 >= len(lines):
        return False
    return "|" in lines[index] and "|" in lines[index + 1] and set(lines[index + 1].strip().replace("|", "").replace(":", "").replace(" ", "")) <= {"-"}


def _markdown_table_block_length(lines, index):
    end = index + 2
    while end < len(lines) and "|" in lines[end] and lines[end].strip():
        end += 1
    return end - index


def _parse_markdown_table(lines):
    rows = []
    for line in lines:
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if cells and not all(set(cell.replace(":", "").replace(" ", "")) <= {"-"} for cell in cells):
            rows.append(cells)
    return rows


def _markdown_table_to_html(lines):
    rows = _parse_markdown_table(lines)
    if not rows:
        return ""
    head = "".join("<th>" + _inline_markdown_to_html(cell) + "</th>" for cell in rows[0])
    body = []
    for row in rows[1:]:
        body.append("<tr>" + "".join("<td>" + _inline_markdown_to_html(cell) + "</td>" for cell in row) + "</tr>")
    return "<table><thead><tr>{}</tr></thead><tbody>{}</tbody></table>".format(head, "".join(body))


def _write_excel(path, sheets):
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    if isinstance(sheets, dict):
        iterable = sheets.items()
    elif isinstance(sheets, list):
        iterable = []
        for index, item in enumerate(sheets, start=1):
            if isinstance(item, dict):
                iterable.append((item.get("name") or ("Sheet" + str(index)), item.get("rows") or []))
            else:
                iterable.append(("Sheet" + str(index), item))
    else:
        raise TypeError("sheets must be a dict or list")

    sheet_count = 0
    row_count = 0
    for raw_name, rows in iterable:
        name = str(raw_name or "Sheet")[:31]
        ws = workbook.create_sheet(title=name)
        sheet_count += 1
        for row in rows or []:
            if isinstance(row, dict):
                values = list(row.values())
            elif isinstance(row, (list, tuple)):
                values = list(row)
            else:
                values = [row]
            ws.append([_cell_value(v) for v in values])
            row_count += 1

    if sheet_count == 0:
        workbook.create_sheet(title="Sheet1")
        sheet_count = 1

    workbook.save(path)
    workbook.close()
    return {"path": path, "sheet_count": sheet_count, "row_count": row_count}
